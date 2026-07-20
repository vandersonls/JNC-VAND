import io
import unicodedata

from flask import Blueprint, request, jsonify, send_file
from flask_login import login_required, current_user
import openpyxl
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet

import db
from auth import perfis_permitidos
from auditoria import registrar
from areas import areas_permitidas

materiais_bp = Blueprint("materiais", __name__)

COLUNAS = ["codigo", "descricao", "fabricante", "bitola", "unidade", "area_id"]


def _filtro_area_sql(alias=""):
    """Monta a cláusula SQL (e parâmetros) que restringe materiais às áreas
    que o usuário logado pode ver. Retorna (sql_ou_vazio, params). Master não
    tem restrição (retorna sem filtro)."""
    permitidas = areas_permitidas(current_user)
    if permitidas is None:
        return "", ()
    prefixo = f"{alias}." if alias else ""
    if not permitidas:
        return f" AND 1=0", ()
    placeholders = ", ".join(["%s"] * len(permitidas))
    return f" AND {prefixo}area_id IN ({placeholders})", tuple(permitidas)


def _validar_area_permitida(area_id):
    """Verifica se o usuário logado pode cadastrar/editar materiais na área
    informada. Master sempre pode."""
    permitidas = areas_permitidas(current_user)
    if permitidas is None:
        return True
    return area_id in permitidas


def _chave_duplicidade(row):
    """Materiais com a mesma descrição + fabricante + bitola (ignorando
    acentos/maiúsculas/espaços) são considerados o mesmo item cadastrado sob
    códigos diferentes."""
    partes = (row["descricao"], row["fabricante"], row["bitola"])
    return tuple(_normalizar(p) for p in partes)


def _marcar_duplicados(rows):
    contagem = {}
    for row in rows:
        chave = _chave_duplicidade(row)
        contagem[chave] = contagem.get(chave, 0) + 1
    for row in rows:
        row["duplicado"] = contagem[_chave_duplicidade(row)] > 1
    return rows


@materiais_bp.get("/api/materiais")
@login_required
def listar_materiais():
    busca = request.args.get("q", "").strip()
    somente_duplicados = request.args.get("somente_duplicados") == "1"
    filtro_area_sql, filtro_area_params = _filtro_area_sql(alias="m")

    base_select = """SELECT m.*, a.nome AS area_nome FROM materiais m
                      LEFT JOIN areas a ON a.id = m.area_id"""

    if busca:
        like = f"%{busca}%"
        rows = db.query_all(
            f"""{base_select}
               WHERE m.ativo = 1 AND (m.codigo LIKE %s OR m.descricao LIKE %s OR m.fabricante LIKE %s)
               {filtro_area_sql}
               ORDER BY m.codigo""",
            (like, like, like, *filtro_area_params),
        )
    else:
        rows = db.query_all(
            f"{base_select} WHERE m.ativo = 1 {filtro_area_sql} ORDER BY m.codigo",
            filtro_area_params,
        )

    if somente_duplicados:
        # A comparação de duplicidade precisa considerar todo o catálogo (dentro
        # das áreas permitidas), não só os resultados já filtrados pela busca.
        todos = rows if not busca else db.query_all(
            f"{base_select} WHERE m.ativo = 1 {filtro_area_sql}", filtro_area_params,
        )
        contagem = {}
        for row in todos:
            chave = _chave_duplicidade(row)
            contagem[chave] = contagem.get(chave, 0) + 1
        rows = [r for r in rows if contagem[_chave_duplicidade(r)] > 1]
        for r in rows:
            r["duplicado"] = True
    else:
        _marcar_duplicados(rows)

    return jsonify(rows)


@materiais_bp.post("/api/materiais")
@perfis_permitidos("master", "administrador")
def criar_material():
    data = request.get_json(force=True) or {}
    faltando = [c for c in COLUNAS if not data.get(c)]
    if faltando:
        return jsonify({"erro": f"Campos obrigatórios: {', '.join(faltando)}"}), 400
    if not _validar_area_permitida(int(data["area_id"])):
        return jsonify({"erro": "Você não tem permissão para cadastrar materiais nesta área"}), 403
    novo_id = db.execute(
        """INSERT INTO materiais (codigo, descricao, fabricante, bitola, unidade, area_id)
           VALUES (%s, %s, %s, %s, %s, %s)""",
        tuple(data[c] for c in COLUNAS),
    )
    registrar("criar", "material", novo_id, f"Criou o material {data['codigo']}", depois=data)
    return jsonify({"id": novo_id}), 201


@materiais_bp.put("/api/materiais/<int:material_id>")
@perfis_permitidos("master", "administrador")
def editar_material(material_id):
    data = request.get_json(force=True) or {}
    faltando = [c for c in COLUNAS if not data.get(c)]
    if faltando:
        return jsonify({"erro": f"Campos obrigatórios: {', '.join(faltando)}"}), 400
    antes = db.query_one("SELECT codigo, descricao, fabricante, bitola, unidade, area_id FROM materiais WHERE id = %s", (material_id,))
    if antes and not _validar_area_permitida(antes["area_id"]):
        return jsonify({"erro": "Você não tem permissão para editar materiais desta área"}), 403
    if not _validar_area_permitida(int(data["area_id"])):
        return jsonify({"erro": "Você não tem permissão para mover o material para esta área"}), 403
    db.execute(
        """UPDATE materiais SET codigo=%s, descricao=%s, fabricante=%s, bitola=%s, unidade=%s, area_id=%s
           WHERE id=%s""",
        (*[data[c] for c in COLUNAS], material_id),
    )
    registrar("editar", "material", material_id, f"Editou o material {data['codigo']}", antes=antes, depois=data)
    return jsonify({"ok": True})


@materiais_bp.delete("/api/materiais/<int:material_id>")
@perfis_permitidos("master", "administrador")
def excluir_material(material_id):
    antes = db.query_one("SELECT codigo, descricao, area_id FROM materiais WHERE id = %s", (material_id,))
    if antes and not _validar_area_permitida(antes["area_id"]):
        return jsonify({"erro": "Você não tem permissão para excluir materiais desta área"}), 403
    db.execute("UPDATE materiais SET ativo = 0 WHERE id = %s", (material_id,))
    if antes:
        registrar("excluir", "material", material_id, f"Excluiu o material {antes['codigo']}", antes=antes)
    return jsonify({"ok": True})


@materiais_bp.post("/api/materiais/excluir-lote")
@perfis_permitidos("master", "administrador")
def excluir_materiais_lote():
    data = request.get_json(force=True) or {}
    ids = data.get("ids") or []
    if not ids:
        return jsonify({"erro": "Nenhum material selecionado"}), 400

    filtro_area_sql, filtro_area_params = _filtro_area_sql()

    excluidos = []
    for i in range(0, len(ids), 1000):
        lote = ids[i:i + 1000]
        placeholders = ", ".join(["%s"] * len(lote))
        antes = db.query_all(
            f"SELECT codigo FROM materiais WHERE id IN ({placeholders}) {filtro_area_sql}",
            (*lote, *filtro_area_params),
        )
        excluidos.extend(r["codigo"] for r in antes)
        db.execute(
            f"UPDATE materiais SET ativo = 0 WHERE id IN ({placeholders}) {filtro_area_sql}",
            (*lote, *filtro_area_params),
        )

    resumo_codigos = ", ".join(excluidos[:20]) + (f" e mais {len(excluidos) - 20}" if len(excluidos) > 20 else "")
    registrar(
        "excluir", "material", None,
        f"Excluiu {len(excluidos)} material(is) em lote: {resumo_codigos}",
        antes={"codigos": excluidos[:200]},
    )
    return jsonify({"excluidos": len(excluidos)})


@materiais_bp.get("/api/materiais/exportar/excel")
@login_required
def exportar_excel():
    filtro_area_sql, filtro_area_params = _filtro_area_sql(alias="m")
    rows = db.query_all(
        f"""SELECT m.*, a.nome AS area_nome FROM materiais m LEFT JOIN areas a ON a.id = m.area_id
            WHERE m.ativo = 1 {filtro_area_sql} ORDER BY m.codigo""",
        filtro_area_params,
    )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Materiais"
    cabecalho = ["Código", "Descrição", "Fabricante", "Bitola", "Unidade", "Área"]
    ws.append(cabecalho)
    for row in rows:
        ws.append([row["codigo"], row["descricao"], row["fabricante"], row["bitola"], row["unidade"], row["area_nome"]])
    for i, _ in enumerate(cabecalho, start=1):
        ws.column_dimensions[get_column_letter(i)].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="materiais.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@materiais_bp.get("/api/materiais/exportar/pdf")
@login_required
def exportar_pdf():
    filtro_area_sql, filtro_area_params = _filtro_area_sql(alias="m")
    rows = db.query_all(
        f"""SELECT m.*, a.nome AS area_nome FROM materiais m LEFT JOIN areas a ON a.id = m.area_id
            WHERE m.ativo = 1 {filtro_area_sql} ORDER BY m.codigo""",
        filtro_area_params,
    )
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elementos = [Paragraph("Lista de Materiais", styles["Title"])]

    dados = [["Código", "Descrição", "Fabricante", "Bitola", "Unidade", "Área"]]
    for row in rows:
        dados.append([row["codigo"], row["descricao"], row["fabricante"] or "", row["bitola"] or "", row["unidade"], row["area_nome"] or ""])

    tabela = Table(dados, repeatRows=1)
    tabela.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f3a5f")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f2f4f7")]),
    ]))
    elementos.append(tabela)
    doc.build(elementos)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="materiais.pdf", mimetype="application/pdf")


def _normalizar(texto):
    """Remove acentos e deixa minúsculo, para comparar cabeçalhos com tolerância a variações."""
    if texto is None:
        return ""
    sem_acento = unicodedata.normalize("NFKD", str(texto)).encode("ascii", "ignore").decode("ascii")
    return sem_acento.strip().lower()


# Cada campo interno aceita várias grafias possíveis encontradas em planilhas reais
SINONIMOS = {
    "codigo": ["codigo"],
    "descricao": ["descricao"],
    "fabricante": ["fabricante", "fabricacao", "fabricado", "marca"],
    "bitola": ["bitola"],
    "unidade": ["unidade", "un", "und", "unid"],
}


def _texto(valor, limite):
    """Converte qualquer valor de célula (Excel às vezes entrega número ou
    data mesmo quando o conteúdo deveria ser texto) para string segura,
    cortada no limite da coluna do banco para nunca quebrar o INSERT."""
    if valor is None:
        return ""
    return str(valor).strip()[:limite]


def _ler_planilha(arquivo):
    """Lê e valida a planilha, devolvendo (total_linhas, ignoradas, linhas_validas).
    Cada item de linhas_validas é (numero_da_linha, codigo, descricao, fabricante, bitola, unidade).
    Lança ValueError com mensagem amigável se algo estiver errado."""
    wb = openpyxl.load_workbook(arquivo, data_only=True)
    ws = wb.active

    linhas = list(ws.iter_rows(values_only=True))
    if not linhas:
        raise ValueError("Planilha vazia")

    cabecalho = [_normalizar(c) for c in linhas[0]]
    idx = {}
    faltando = []
    for campo, variantes in SINONIMOS.items():
        posicao = next((cabecalho.index(v) for v in variantes if v in cabecalho), None)
        if posicao is None:
            faltando.append(campo)
        else:
            idx[campo] = posicao
    if faltando:
        raise ValueError(f"Colunas não encontradas na planilha: {', '.join(faltando)}")

    total_linhas = len(linhas) - 1
    ignoradas = 0
    linhas_validas = []
    for n, linha in enumerate(linhas[1:], start=2):
        codigo = linha[idx["codigo"]]
        if not codigo:
            ignoradas += 1
            continue
        linhas_validas.append((
            n,
            _texto(codigo, 50),
            _texto(linha[idx["descricao"]], 500),
            _texto(linha[idx["fabricante"]], 150),
            _texto(linha[idx["bitola"]], 50),
            _texto(linha[idx["unidade"]], 20),
        ))
    return total_linhas, ignoradas, linhas_validas


def _agrupar_duplicados(linhas_validas):
    """Agrupa linhas pelo código e sinaliza quais grupos têm dados
    divergentes entre si (não basta o código repetir - descrição, fabricante
    e bitola também precisam ser comparados para saber se é uma duplicidade
    inofensiva ou um conflito real)."""
    por_codigo = {}
    for n, codigo, descricao, fabricante, bitola, unidade in linhas_validas:
        por_codigo.setdefault(codigo, []).append({
            "linha": n, "descricao": descricao, "fabricante": fabricante, "bitola": bitola, "unidade": unidade,
        })

    duplicados = []
    for codigo, ocorrencias in por_codigo.items():
        if len(ocorrencias) < 2:
            continue
        campos_distintos = {(o["descricao"], o["fabricante"], o["bitola"], o["unidade"]) for o in ocorrencias}
        duplicados.append({
            "codigo": codigo,
            "linhas": [o["linha"] for o in ocorrencias],
            "conflito": len(campos_distintos) > 1,
            "ocorrencias": ocorrencias,
        })
    duplicados.sort(key=lambda d: (not d["conflito"], d["codigo"]))
    return duplicados


@materiais_bp.post("/api/materiais/importar/excel/analisar")
@perfis_permitidos("master", "administrador")
def analisar_importacao_excel():
    arquivo = request.files.get("arquivo")
    if not arquivo:
        return jsonify({"erro": "Nenhum arquivo enviado"}), 400
    try:
        total_linhas, ignoradas, linhas_validas = _ler_planilha(arquivo)
    except ValueError as e:
        return jsonify({"erro": str(e)}), 400

    duplicados = _agrupar_duplicados(linhas_validas)
    codigos_unicos = len({c for _, c, *_ in linhas_validas})

    return jsonify({
        "total_linhas": total_linhas,
        "ignoradas": ignoradas,
        "codigos_unicos": codigos_unicos,
        "duplicados": duplicados,
    })


@materiais_bp.post("/api/materiais/importar/excel")
@perfis_permitidos("master", "administrador")
def importar_excel():
    arquivo = request.files.get("arquivo")
    if not arquivo:
        return jsonify({"erro": "Nenhum arquivo enviado"}), 400
    area_id = request.form.get("area_id", type=int)
    if not area_id:
        return jsonify({"erro": "Selecione a área para a qual esses materiais serão importados"}), 400
    if not _validar_area_permitida(area_id):
        return jsonify({"erro": "Você não tem permissão para importar materiais nesta área"}), 403
    # "manter" (padrão): a última ocorrência de cada código repetido prevalece.
    # "excluir": nenhuma linha com código repetido é importada (nem a primeira).
    modo_duplicados = request.form.get("duplicados", "manter")

    try:
        total_linhas, ignoradas, linhas_validas = _ler_planilha(arquivo)
    except ValueError as e:
        return jsonify({"erro": str(e)}), 400

    duplicados_excluidos = 0
    if modo_duplicados == "excluir":
        contagem = {}
        for _, codigo, *_ in linhas_validas:
            contagem[codigo] = contagem.get(codigo, 0) + 1
        antes = len(linhas_validas)
        linhas_validas = [item for item in linhas_validas if contagem[item[1]] == 1]
        duplicados_excluidos = antes - len(linhas_validas)

    linhas_dados = [(codigo, descricao, fabricante, bitola, unidade, area_id)
                     for _, codigo, descricao, fabricante, bitola, unidade in linhas_validas]

    # Descobre de uma vez quais códigos já existem, em vez de 1 SELECT por linha
    # (essencial para não estourar o timeout do servidor em planilhas grandes).
    codigos_existentes = set()
    todos_codigos = [c[0] for c in linhas_dados]
    for i in range(0, len(todos_codigos), 1000):
        lote = todos_codigos[i:i + 1000]
        if not lote:
            continue
        placeholders = ", ".join(["%s"] * len(lote))
        rows = db.query_all(f"SELECT codigo FROM materiais WHERE codigo IN ({placeholders})", tuple(lote))
        codigos_existentes.update(r["codigo"] for r in rows)

    inseridos, atualizados = 0, 0
    ja_vistos = set()
    for codigo, *_ in linhas_dados:
        if codigo in codigos_existentes or codigo in ja_vistos:
            atualizados += 1
        else:
            inseridos += 1
        ja_vistos.add(codigo)

    # Grava tudo em lotes (executemany), bem mais rápido que uma query por linha
    sql_upsert = """
        INSERT INTO materiais (codigo, descricao, fabricante, bitola, unidade, area_id, ativo)
        VALUES (%s, %s, %s, %s, %s, %s, 1)
        ON DUPLICATE KEY UPDATE
            descricao = VALUES(descricao), fabricante = VALUES(fabricante),
            bitola = VALUES(bitola), unidade = VALUES(unidade), area_id = VALUES(area_id), ativo = 1
    """
    for i in range(0, len(linhas_dados), 500):
        lote = linhas_dados[i:i + 500]
        if lote:
            db.execute_many(sql_upsert, lote)

    registrar(
        "importar", "material", None,
        f"Importou planilha '{arquivo.filename}' (duplicados: {modo_duplicados}): {total_linhas} linha(s) lida(s), "
        f"{inseridos} novos, {atualizados} atualizados, {ignoradas} ignorada(s), "
        f"{duplicados_excluidos} excluída(s) por duplicidade",
    )
    return jsonify({
        "total_linhas": total_linhas,
        "inseridos": inseridos,
        "atualizados": atualizados,
        "ignoradas": ignoradas,
        "duplicados_excluidos": duplicados_excluidos,
        "erros": [],
    })
