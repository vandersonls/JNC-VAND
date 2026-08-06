import io
import ipaddress
import os
import re
import socket
import urllib.request
from copy import copy
from urllib.parse import urlparse

import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from flask import Blueprint, request, jsonify, send_file
from flask_login import login_required
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image as PdfImage

import db
from areas import projeto_permitido, projeto_da_lista

relatorios_bp = Blueprint("relatorios", __name__)


def _url_segura_para_baixar(url):
    """Proteção contra SSRF: só permite http/https apontando para endereços
    públicos. Bloqueia esquemas perigosos (file://, etc.) e qualquer host que
    resolva para IP interno/privado (loopback, rede local, metadados da nuvem)."""
    try:
        p = urlparse(url)
    except Exception:
        return False
    if p.scheme not in ("http", "https") or not p.hostname:
        return False
    porta = p.port or (443 if p.scheme == "https" else 80)
    try:
        enderecos = socket.getaddrinfo(p.hostname, porta)
    except Exception:
        return False
    for info in enderecos:
        try:
            ip = ipaddress.ip_address(info[4][0])
        except ValueError:
            return False
        if (ip.is_private or ip.is_loopback or ip.is_link_local
                or ip.is_reserved or ip.is_multicast or ip.is_unspecified):
            return False
    return True


def _baixar_imagem(url):
    """Baixa uma imagem de logo por URL. Nunca lança exceção - um link de
    logo quebrado, lento ou não permitido não pode impedir a geração do
    relatório. Limita o tamanho lido para não estourar a memória do servidor."""
    if not url or not _url_segura_para_baixar(url):
        return None
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=5) as resp:
            return resp.read(5 * 1024 * 1024)  # no máximo 5 MB (logo real é bem menor)
    except Exception:
        return None


def _cabecalho_pdf_com_logos(empresa_nome, empresa_logo_url, cliente_nome, cliente_logo_url,
                              projeto_nome, nome_aba, revisao, data_texto):
    """Monta o bloco de cabeçalho padrão (empresa+logo | cliente+logo, depois
    projeto/aba/revisão) usado nos relatórios de Lista por Desenho, Lista PQ
    e Lista de Compras."""
    estilo_nome = ParagraphStyle("cab_nome", fontSize=10, fontName="Helvetica-Bold", alignment=1)
    estilo_rotulo = ParagraphStyle("cab_rotulo", fontSize=7.5, textColor=colors.grey, alignment=1)

    def _celula_logo(nome, url, rotulo):
        img_bytes = _baixar_imagem(url)
        partes = [Paragraph(rotulo, estilo_rotulo)]
        if img_bytes:
            try:
                img = PdfImage(io.BytesIO(img_bytes), width=2.6 * cm, height=2.6 * cm, kind="proportional")
                img.hAlign = "CENTER"
                partes.append(img)
            except Exception:
                pass
        partes.append(Paragraph(nome or "-", estilo_nome))
        return partes

    tabela_logos = Table(
        [[_celula_logo(empresa_nome, empresa_logo_url, "EMPRESA"), _celula_logo(cliente_nome, cliente_logo_url, "CLIENTE")]],
        colWidths=[13.5 * cm, 13.5 * cm],
    )
    tabela_logos.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.5, colors.grey),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))

    titulo_estilo = ParagraphStyle("titulo", fontSize=15, leading=18, alignment=1, spaceAfter=4,
                                    spaceBefore=10, fontName="Helvetica-Bold")
    subtitulo_estilo = ParagraphStyle("subtitulo", fontSize=11, leading=14, alignment=1, spaceAfter=4, fontName="Helvetica-Bold")
    ref_estilo = ParagraphStyle("ref", fontSize=9, leading=12, alignment=1, spaceAfter=10)

    return [
        tabela_logos,
        Paragraph(projeto_nome, titulo_estilo),
        Paragraph(nome_aba, subtitulo_estilo),
        Paragraph(f"Revisão: {revisao} &nbsp;&nbsp;|&nbsp;&nbsp; Data: {data_texto}", ref_estilo),
    ]


def _cabecalho_excel_com_logos(ws, largura_total, empresa_nome, empresa_logo_url, cliente_nome, cliente_logo_url,
                                projeto_nome, nome_aba, revisao, data_texto, linha_inicial=1):
    """Escreve o cabeçalho padrão (com logos, se disponíveis) numa planilha
    e devolve a próxima linha livre."""
    centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    linha = linha_inicial

    ws.row_dimensions[linha].height = 48
    for offset, (nome, url, rotulo) in enumerate([
        (empresa_nome, empresa_logo_url, "EMPRESA"), (cliente_nome, cliente_logo_url, "CLIENTE"),
    ]):
        col_ini = 1 + offset * (largura_total // 2)
        col_fim = col_ini + (largura_total // 2) - 1
        ws.merge_cells(start_row=linha, start_column=col_ini, end_row=linha, end_column=col_fim)
        cel = ws.cell(row=linha, column=col_ini, value=f"{rotulo}: {nome or '-'}")
        cel.font, cel.alignment = Font(bold=True, size=10), centro
        img_bytes = _baixar_imagem(url)
        if img_bytes:
            try:
                img = ExcelImage(io.BytesIO(img_bytes))
                img.width, img.height = 60, 60
                ws.add_image(img, f"{get_column_letter(col_ini)}{linha}")
            except Exception:
                pass
    linha += 1

    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=largura_total)
    cel = ws.cell(row=linha, column=1, value=projeto_nome)
    cel.font, cel.alignment = Font(bold=True, size=13), centro
    linha += 1

    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=largura_total)
    cel = ws.cell(row=linha, column=1, value=nome_aba)
    cel.font, cel.alignment = Font(bold=True, size=10), centro
    linha += 1

    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=largura_total)
    cel = ws.cell(row=linha, column=1, value=f"Revisão: {revisao}    |    Data: {data_texto}")
    cel.font, cel.alignment = Font(size=9), centro
    linha += 2
    return linha

CEL_ESTILO = ParagraphStyle("cel", fontSize=8, leading=10)
CEL_ESTILO_CABECALHO = ParagraphStyle("cel_cab", fontSize=8, leading=10, textColor=colors.white, fontName="Helvetica-Bold")


def _tabela_quebravel(dados, col_widths, alinhar_direita=None, cor_cabecalho=None, cor_texto_cabecalho=None):
    """Monta uma Table do reportlab em que toda célula é um Paragraph,
    para que o texto quebre linha dentro da largura da coluna em vez de
    transbordar/sobrepor o texto vizinho."""
    alinhar_direita = alinhar_direita or set()
    cor_cabecalho = cor_cabecalho or colors.HexColor("#1f3a5f")
    estilo_cabecalho = (
        CEL_ESTILO_CABECALHO if cor_texto_cabecalho is None
        else ParagraphStyle("cel_cab_custom", fontSize=8, leading=10, textColor=cor_texto_cabecalho, fontName="Helvetica-Bold")
    )
    linhas = []
    for i, linha in enumerate(dados):
        estilo = estilo_cabecalho if i == 0 else CEL_ESTILO
        linhas.append([Paragraph(str(v), estilo) for v in linha])
    tabela = Table(linhas, colWidths=col_widths, repeatRows=1)
    estilo_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), cor_cabecalho),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f2f4f7")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    for col in alinhar_direita:
        estilo_cmds.append(("ALIGN", (col, 0), (col, -1), "CENTER"))
    tabela.setStyle(TableStyle(estilo_cmds))
    return tabela


def _carregar_contexto(lista_id, versao_id=None):
    lista = db.query_one(
        """SELECT ld.*, p.codigo AS projeto_codigo, p.nome AS projeto_nome,
                  c.razao_social AS cliente_nome, c.logo_url AS cliente_logo_url
           FROM listas_desenho ld
           JOIN projetos p ON p.id = ld.projeto_id
           LEFT JOIN clientes c ON c.id = p.cliente_id
           WHERE ld.id = %s""",
        (lista_id,),
    )
    if not lista:
        return None

    alvo_versao_id = versao_id or lista["versao_atual_id"]
    if not alvo_versao_id:
        return {"lista": lista, "versao": None, "itens": [], "historico": []}

    versao = db.query_one(
        """SELECT v.*, u.nome AS criado_por_nome FROM lista_desenho_versoes v
           LEFT JOIN usuarios u ON u.id = v.criado_por WHERE v.id = %s""",
        (alvo_versao_id,),
    )
    itens = db.query_all(
        """SELECT i.material_id, i.quantidade, i.observacao, m.codigo, m.descricao, m.fabricante, m.bitola, m.unidade
           FROM lista_desenho_itens i JOIN materiais m ON m.id = i.material_id
           WHERE i.versao_id = %s ORDER BY m.codigo""",
        (alvo_versao_id,),
    )
    historico = db.query_all(
        """SELECT v.versao, v.status, v.tipo_emissao, v.observacoes, v.criado_em, u.nome AS criado_por_nome
           FROM lista_desenho_versoes v LEFT JOIN usuarios u ON u.id = v.criado_por
           WHERE v.lista_desenho_id = %s ORDER BY v.versao""",
        (lista_id,),
    )
    config = {c["chave"]: c["valor"] for c in db.query_all("SELECT chave, valor FROM configuracoes")}

    qtd_anterior_por_material = _quantidades_versao_anterior(lista_id, versao)
    for item in itens:
        item["quantidade_anterior"] = qtd_anterior_por_material.get(item["material_id"], 0)

    return {"lista": lista, "versao": versao, "itens": itens, "historico": historico, "config": config}


def _quantidades_versao_anterior(lista_id, versao_atual):
    """Quantidade de cada material na última versão EMITIDA anterior à
    atual (por número de revisão) - usada na coluna QUANT. ANTERIOR do
    relatório, pra mostrar de imediato o que mudou de uma revisão pra outra."""
    if not versao_atual:
        return {}
    anterior = db.query_one(
        """SELECT id FROM lista_desenho_versoes
           WHERE lista_desenho_id = %s AND status = 'salvo' AND versao < %s
           ORDER BY versao DESC LIMIT 1""",
        (lista_id, versao_atual["versao"]),
    )
    if not anterior:
        return {}
    linhas = db.query_all(
        "SELECT material_id, quantidade FROM lista_desenho_itens WHERE versao_id = %s",
        (anterior["id"],),
    )
    return {l["material_id"]: l["quantidade"] for l in linhas}


def _doc_referencia(lista):
    return f"{lista['projeto_codigo']}-{lista['numero_desenho']}"


def _carregar_contexto_projeto(projeto_id):
    """Reúne, para cada Lista por Desenho do projeto, sempre a sua última versão salva."""
    if not projeto_permitido(projeto_id):
        return None
    projeto = db.query_one(
        """SELECT p.*, c.razao_social AS cliente_nome
           FROM projetos p LEFT JOIN clientes c ON c.id = p.cliente_id
           WHERE p.id = %s""",
        (projeto_id,),
    )
    if not projeto:
        return None

    listas = db.query_all(
        "SELECT * FROM listas_desenho WHERE projeto_id = %s ORDER BY numero_desenho",
        (projeto_id,),
    )

    desenhos = []
    consolidado = {}
    for lista in listas:
        versao, itens = None, []
        if lista["versao_atual_id"]:
            versao = db.query_one(
                """SELECT v.*, u.nome AS criado_por_nome FROM lista_desenho_versoes v
                   LEFT JOIN usuarios u ON u.id = v.criado_por WHERE v.id = %s""",
                (lista["versao_atual_id"],),
            )
            itens = db.query_all(
                """SELECT i.quantidade, i.observacao, m.codigo, m.descricao, m.fabricante, m.bitola, m.unidade
                   FROM lista_desenho_itens i JOIN materiais m ON m.id = i.material_id
                   WHERE i.versao_id = %s ORDER BY m.codigo""",
                (lista["versao_atual_id"],),
            )
        desenhos.append({"lista": lista, "versao": versao, "itens": itens})

        for item in itens:
            chave = item["codigo"]
            if chave not in consolidado:
                consolidado[chave] = {
                    "codigo": item["codigo"], "descricao": item["descricao"], "fabricante": item["fabricante"],
                    "bitola": item["bitola"], "unidade": item["unidade"], "quantidade": 0, "desenhos": set(),
                }
            consolidado[chave]["quantidade"] += float(item["quantidade"])
            consolidado[chave]["desenhos"].add(lista["numero_desenho"])

    consolidado_lista = sorted(consolidado.values(), key=lambda x: x["codigo"])
    config = {c["chave"]: c["valor"] for c in db.query_all("SELECT chave, valor FROM configuracoes")}

    return {"projeto": projeto, "desenhos": desenhos, "consolidado": consolidado_lista, "config": config}


def _assinatura_curta(lista, campo_nome, campo_sigla):
    """Sigla se houver, senão o nome completo, senão traço - usado nas
    colunas estreitas (Por/Ver./Apr./Aut.) do histórico de revisões."""
    return lista.get(campo_sigla) or lista.get(campo_nome) or "-"


def _rev_exibicao(lista, versao):
    """Rev. preenchida à mão no cabeçalho tem prioridade (o operador confere
    com a última revisão registrada); sem isso, cai no número de versão
    automático do sistema."""
    if lista.get("rev_manual") is not None:
        return lista["rev_manual"]
    return versao["versao"] if versao else "-"


def _data_emissao_exibicao(lista, versao):
    """Mesma lógica da Rev.: data preenchida à mão tem prioridade (permite
    retroagir a data de documentos migrados), senão usa a data real de
    criação da versão no sistema."""
    manual = lista.get("data_emissao_manual")
    if manual:
        return manual.strftime("%d/%m/%Y") if hasattr(manual, "strftime") else str(manual)
    return versao["criado_em"].strftime("%d/%m/%Y") if versao else "-"


# Fonte e tamanhos usados no carimbo padrão de documentos de engenharia
# (modelo em anexo do cliente) - bem mais compactos que o resto do sistema.
# =========================================================
# LISTA POR DESENHO - preenchimento do molde exato do cliente
# =========================================================
# O arquivo relatorio_templates/lista_por_desenho.xlsx é o modelo enviado
# pelo cliente e NÃO pode ser alterado (símbolos, logos, fontes, bordas,
# layout). Só escrevemos valores nas células de dados - toda a formatação
# é a que já vem no próprio arquivo.
TEMPLATE_LISTA_DESENHO = os.path.join(os.path.dirname(__file__), "relatorio_templates", "lista_por_desenho.xlsx")

# aba 1 = "MMITT-ED-LM-..." (itens), aba 0 = "Capa " - mesmas coordenadas de
# cabeçalho nas duas, exceto onde a largura da aba muda a coluna do bloco
# Nº Cliente/Nº Projetista/Projeto (W na aba de itens, V na Capa).
_CAMPOS_CABECALHO_ITENS = {
    "projeto": "W2", "subtitulo": "B4", "area": "B5", "disciplina": "B6", "titulo": "B7",
    "numero_cliente": "W4", "numero_projetista": "W7", "rev": "AG7",
}
_CAMPOS_CABECALHO_CAPA = {
    "projeto": "V2", "subtitulo": "B4", "area": "B5", "disciplina": "B6", "titulo": "B7",
    "numero_cliente": "V4", "numero_projetista": "V7", "rev": "AG7",
}

# (nome_campo, coluna_inicial, coluna_final) de cada linha de item na tabela
# de materiais - a mescla de cada campo já vem pronta no molde até a linha 42.
_ITEM_COLUNAS = [
    ("item", 2, 2), ("codigo", 3, 4), ("descricao", 5, 15), ("referencia", 16, 22),
    ("complemento", 23, 27), ("unidade", 28, 29), ("quant_atual", 30, 32), ("quant_anterior", 33, 34),
]
_ITEM_LINHA_INICIAL = 11
_ITEM_LINHA_FINAL_MOLDE = 42
_ITEM_LINHA_ESTILO = 20  # linha "do meio" usada como fonte de estilo ao precisar de mais linhas que o molde

_REV_COLUNAS = [
    ("rev", 2, 3), ("te", 4, 5), ("descricao", 6, 18), ("por", 19, 21),
    ("ver", 22, 24), ("apr", 25, 27), ("aut", 28, 30), ("data", 31, 36),
]
_REV_LINHA_INICIAL = 14
_REV_LINHA_FINAL_MOLDE = 31
_REV_LINHA_ESTILO = 20


def _titulo_aba_valido(texto):
    """Nome de aba do Excel não aceita : \\ / ? * [ ] nem mais de 31 caracteres."""
    texto = re.sub(r'[:\\/?*\[\]]', "-", texto or "")
    return texto[:31] or "Lista"


def _duplicar_linha_estilo(ws, linha_origem, linha_destino, colunas):
    """Copia formatação (fonte/borda/preenchimento/alinhamento) e mesclagens
    de uma linha do molde pra uma linha nova - usado quando a lista tem mais
    itens/revisões do que as linhas já prontas no arquivo do cliente."""
    ws.row_dimensions[linha_destino].height = ws.row_dimensions[linha_origem].height
    for _, col_ini, col_fim in colunas:
        for col in range(col_ini, col_fim + 1):
            origem = ws.cell(row=linha_origem, column=col)
            destino = ws.cell(row=linha_destino, column=col)
            destino.font = copy(origem.font)
            destino.border = copy(origem.border)
            destino.fill = copy(origem.fill)
            destino.alignment = copy(origem.alignment)
            destino.number_format = origem.number_format
        if col_fim > col_ini:
            ws.merge_cells(start_row=linha_destino, start_column=col_ini, end_row=linha_destino, end_column=col_fim)


def _escrever_linha_grade(ws, linha, colunas, valores):
    for nome, col_ini, _ in colunas:
        if nome in valores:
            ws.cell(row=linha, column=col_ini, value=valores[nome])


def _definir_com_quebra(ws, coord, valor, tamanho_min=None, limite_caracteres=40):
    """Escreve o valor mantendo o resto do estilo da célula, mas ligando
    quebra de linha automática - o molde vem sem isso, então texto mais
    longo que o esperado (ex.: nome de projeto grande) vazava pra fora da
    caixa em vez de quebrar dentro dela. Se mesmo quebrando o texto for
    comprido demais, reduz um pouco a fonte em vez de deixar cortado."""
    cel = ws[coord]
    cel.value = valor
    alin = cel.alignment
    cel.alignment = Alignment(horizontal=alin.horizontal, vertical=alin.vertical, wrap_text=True)
    if tamanho_min is not None and valor and len(str(valor)) > limite_caracteres:
        fonte = cel.font
        novo_tamanho = max(tamanho_min, fonte.size - 2)
        cel.font = Font(name=fonte.name, size=novo_tamanho, bold=fonte.bold, italic=fonte.italic, color=fonte.color)


def _preencher_cabecalho_molde(ws, campos, lista, versao):
    _definir_com_quebra(ws, campos["projeto"], lista["projeto_nome"], tamanho_min=9)
    _definir_com_quebra(ws, campos["subtitulo"], lista.get("subtitulo") or "")
    _definir_com_quebra(ws, campos["area"], lista.get("area_titulo") or "")
    _definir_com_quebra(ws, campos["disciplina"], lista.get("disciplina") or "")
    _definir_com_quebra(ws, campos["titulo"], lista.get("titulo") or "")
    _definir_com_quebra(ws, campos["numero_cliente"], lista.get("numero_cliente") or "")
    _definir_com_quebra(ws, campos["numero_projetista"], lista.get("numero_fornecedor") or "")
    _definir_com_quebra(ws, campos["rev"], _rev_exibicao(lista, versao))


def _preencher_itens_molde(ws, itens):
    linha = _ITEM_LINHA_INICIAL
    for idx, item in enumerate(itens, start=1):
        if linha > _ITEM_LINHA_FINAL_MOLDE:
            _duplicar_linha_estilo(ws, _ITEM_LINHA_ESTILO, linha, _ITEM_COLUNAS)
        _escrever_linha_grade(ws, linha, _ITEM_COLUNAS, {
            "item": idx, "codigo": item["codigo"], "descricao": item["descricao"],
            "referencia": item["fabricante"] or "", "complemento": item["bitola"] or "",
            "unidade": item["unidade"], "quant_atual": float(item["quantidade"]),
            "quant_anterior": float(item["quantidade_anterior"]),
        })
        linha += 1
    return max(_ITEM_LINHA_FINAL_MOLDE, linha - 1)


def _preencher_revisoes_molde(ws, lista, versao, historico):
    linha = _REV_LINHA_INICIAL
    for v in (h for h in historico if h["status"] == "salvo"):
        if linha > _REV_LINHA_FINAL_MOLDE:
            _duplicar_linha_estilo(ws, _REV_LINHA_ESTILO, linha, _REV_COLUNAS)
        eh_atual = versao is not None and v["versao"] == versao["versao"]
        data_txt = _data_emissao_exibicao(lista, versao) if eh_atual else v["criado_em"].strftime("%d/%m/%Y")
        _escrever_linha_grade(ws, linha, _REV_COLUNAS, {
            "rev": v["versao"], "te": v["tipo_emissao"] or "-", "descricao": v["observacoes"] or "-",
            "por": _assinatura_curta(lista, "elaborador_nome", "elaborador_sigla"),
            "ver": _assinatura_curta(lista, "verificador_nome", "verificador_sigla"),
            "apr": _assinatura_curta(lista, "aprovador_nome", "aprovador_sigla"),
            "aut": _assinatura_curta(lista, "autorizado_nome", "autorizado_sigla"),
            "data": data_txt,
        })
        linha += 1
    return max(_REV_LINHA_FINAL_MOLDE, linha - 1)


# Coluna mais à direita usada em cada aba, conforme a área de impressão
# original do molde (B2:AH42 na aba de itens, B2:AJ32 na Capa).
_ITEM_COL_DIREITA = 34  # AH
_REV_COL_DIREITA = 36  # AJ


def _ajustar_area_impressao(ws, ultima_linha, col_direita):
    """O molde vem com área e escala de impressão fixas pro tamanho original
    (32 itens / 18 revisões). Se a lista tiver mais linhas que isso, a área
    de impressão precisa crescer junto - senão as linhas extras existem na
    planilha mas não aparecem ao imprimir/exportar (ficam "cortadas"). Troca
    a escala fixa por "encolher até a largura da página" pra nenhuma coluna
    nunca ficar cortada horizontalmente, não importa o conteúdo."""
    ws.print_area = f"B2:{get_column_letter(col_direita)}{ultima_linha}"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def _preencher_molde_lista(lista, versao, itens, historico):
    """Abre o molde do cliente e devolve o Workbook com os dados da lista
    preenchidos nas células certas - sem tocar em nenhuma formatação."""
    wb = openpyxl.load_workbook(TEMPLATE_LISTA_DESENHO)
    ws_capa, ws_itens = wb.worksheets[0], wb.worksheets[1]

    _preencher_cabecalho_molde(ws_itens, _CAMPOS_CABECALHO_ITENS, lista, versao)
    _preencher_cabecalho_molde(ws_capa, _CAMPOS_CABECALHO_CAPA, lista, versao)
    _definir_com_quebra(ws_itens, "B9", f"DESENHO DE REFERÊNCIA : {lista['numero_desenho']}")

    ultima_linha_itens = _preencher_itens_molde(ws_itens, itens)
    ultima_linha_rev = _preencher_revisoes_molde(ws_capa, lista, versao, historico)
    _ajustar_area_impressao(ws_itens, ultima_linha_itens, _ITEM_COL_DIREITA)
    _ajustar_area_impressao(ws_capa, ultima_linha_rev, _REV_COL_DIREITA)

    ws_itens.title = _titulo_aba_valido(lista["numero_desenho"])
    return wb


@relatorios_bp.get("/api/listas/<int:lista_id>/relatorio/excel")
@login_required
def relatorio_excel(lista_id):
    pid = projeto_da_lista(lista_id)
    if pid is None or not projeto_permitido(pid):
        return jsonify({"erro": "Sem permissão para este relatório"}), 403
    ctx = _carregar_contexto(lista_id, request.args.get("versao_id", type=int))
    if not ctx:
        return jsonify({"erro": "Lista não encontrada"}), 404
    lista, versao, itens, historico = ctx["lista"], ctx["versao"], ctx["itens"], ctx["historico"]

    wb = _preencher_molde_lista(lista, versao, itens, historico)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nome_arquivo = f"lista_{lista['numero_desenho']}_rev{versao['versao'] if versao else 0}.xlsx"
    return send_file(buf, as_attachment=True, download_name=nome_arquivo,
                      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@relatorios_bp.get("/api/projetos/<int:projeto_id>/relatorio/excel")
@login_required
def relatorio_projeto_excel(projeto_id):
    ctx = _carregar_contexto_projeto(projeto_id)
    if not ctx:
        return jsonify({"erro": "Projeto não encontrado"}), 404
    projeto, desenhos, consolidado = ctx["projeto"], ctx["desenhos"], ctx["consolidado"]
    empresa = ctx["config"].get("nome_empresa", "")

    wb = openpyxl.Workbook()
    fino = Side(style="thin", color="000000")
    borda = Border(left=fino, right=fino, top=fino, bottom=fino)
    centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    esquerda = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def cabecalho_pagina(ws, largura, subtitulo):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=largura)
        c = ws.cell(row=1, column=1, value=f"{projeto['codigo']} — {projeto['nome']}")
        c.font, c.alignment = Font(bold=True, size=13), centro
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=largura)
        c = ws.cell(row=2, column=1, value=subtitulo)
        c.font, c.alignment = Font(bold=True, size=10), centro
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=largura)
        c = ws.cell(row=3, column=1, value=f"Cliente: {projeto['cliente_nome'] or '-'}    |    Empresa: {empresa}")
        c.font, c.alignment = Font(size=9), centro

    def tabela_com_cabecalho(ws, linha, largura, titulos):
        for col, titulo in enumerate(titulos, start=1):
            cel = ws.cell(row=linha, column=col, value=titulo)
            cel.font = Font(bold=True, size=10, color="FFFFFF")
            cel.fill = openpyxl.styles.PatternFill("solid", fgColor="1f3a5f")
            cel.alignment, cel.border = centro, borda
        return linha + 1

    ws_resumo = wb.active
    ws_resumo.title = "Resumo Consolidado"
    cabecalho_pagina(ws_resumo, 7, "RESUMO CONSOLIDADO DE MATERIAIS (todas as listas, última versão)")
    linha = tabela_com_cabecalho(ws_resumo, 5, 7, ["Código", "Descrição", "Fabricante", "Bitola", "Quantidade", "Unidade", "Desenhos"])
    for item in consolidado:
        valores = [item["codigo"], item["descricao"], item["fabricante"] or "", item["bitola"] or "",
                   item["quantidade"], item["unidade"], ", ".join(sorted(item["desenhos"]))]
        for col, valor in enumerate(valores, start=1):
            cel = ws_resumo.cell(row=linha, column=col, value=valor)
            cel.border = borda
            cel.alignment = esquerda if col in (2, 7) else centro
        linha += 1
    if not consolidado:
        ws_resumo.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=7)
        ws_resumo.cell(row=linha, column=1, value="Nenhum material cadastrado em nenhuma lista deste projeto.").alignment = centro
    for i, w in enumerate([14, 34, 20, 12, 12, 10, 20], start=1):
        ws_resumo.column_dimensions[get_column_letter(i)].width = w

    for d in desenhos:
        lista, versao, itens = d["lista"], d["versao"], d["itens"]
        ws = wb.create_sheet(title=lista["numero_desenho"][:31])
        cabecalho_pagina(ws, 6, f"Desenho {lista['numero_desenho']} — {lista['titulo'] or ''}")
        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=6)
        ws.cell(row=4, column=1, value=f"Revisão: {versao['versao'] if versao else '-'}    |    Nº Cliente: {lista['numero_cliente'] or '-'}    |    Nº Fornecedor: {lista['numero_fornecedor'] or '-'}").alignment = centro
        linha = tabela_com_cabecalho(ws, 6, 6, ["Item", "Código", "Descrição", "Fabricante", "Bitola", "Quantidade"])
        for idx, item in enumerate(itens, start=1):
            valores = [idx, item["codigo"], item["descricao"], item["fabricante"] or "", item["bitola"] or "",
                       float(item["quantidade"])]
            for col, valor in enumerate(valores, start=1):
                cel = ws.cell(row=linha, column=col, value=valor)
                cel.border = borda
                cel.alignment = esquerda if col == 3 else centro
            linha += 1
        if not itens:
            ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=6)
            ws.cell(row=linha, column=1, value="Nenhum material nesta versão.").alignment = centro
        for i, w in enumerate([6, 14, 34, 20, 12, 12], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f"relatorio_projeto_{projeto['codigo']}.xlsx",
                      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@relatorios_bp.get("/api/projetos/<int:projeto_id>/relatorio/pdf")
@login_required
def relatorio_projeto_pdf(projeto_id):
    ctx = _carregar_contexto_projeto(projeto_id)
    if not ctx:
        return jsonify({"erro": "Projeto não encontrado"}), 404
    projeto, desenhos, consolidado = ctx["projeto"], ctx["desenhos"], ctx["consolidado"]
    empresa = ctx["config"].get("nome_empresa", "")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=1.2 * cm, bottomMargin=1.2 * cm,
                             leftMargin=1.2 * cm, rightMargin=1.2 * cm)

    titulo_estilo = ParagraphStyle("titulo", fontSize=15, leading=18, alignment=1, spaceAfter=8, fontName="Helvetica-Bold")
    subtitulo_estilo = ParagraphStyle("subtitulo", fontSize=11, leading=14, alignment=1, spaceAfter=4, fontName="Helvetica-Bold")
    ref_estilo = ParagraphStyle("ref", fontSize=9, leading=12, alignment=1, spaceAfter=10)
    secao_estilo = ParagraphStyle("secao", fontSize=12, fontName="Helvetica-Bold", spaceBefore=14, spaceAfter=4,
                                   textColor=colors.HexColor("#1f3a5f"))
    sub_estilo = ParagraphStyle("sub", fontSize=9, spaceAfter=6)

    elementos = [
        Paragraph(f"{projeto['codigo']} — {projeto['nome']}", titulo_estilo),
        Paragraph("RELATÓRIO FINAL DO PROJETO — CONSOLIDADO DE MATERIAIS", subtitulo_estilo),
        Paragraph(f"Cliente: {projeto['cliente_nome'] or '-'} &nbsp;&nbsp;|&nbsp;&nbsp; Empresa: {empresa} &nbsp;&nbsp;|&nbsp;&nbsp; {len(desenhos)} lista(s) por desenho", ref_estilo),
    ]

    elementos.append(Paragraph("RESUMO CONSOLIDADO (última versão de cada desenho)", secao_estilo))
    dados_consolidado = [["Código", "Descrição", "Fabricante", "Bitola", "Quantidade", "Unidade", "Desenhos"]]
    for item in consolidado:
        dados_consolidado.append([item["codigo"], item["descricao"], item["fabricante"] or "-", item["bitola"] or "-",
                                   f"{item['quantidade']:g}", item["unidade"], ", ".join(sorted(item["desenhos"]))])
    if not consolidado:
        dados_consolidado.append(["-", "Nenhum material cadastrado em nenhuma lista deste projeto.", "-", "-", "-", "-", "-"])
    tabela_consolidado = _tabela_quebravel(
        dados_consolidado,
        col_widths=[2.3 * cm, 7 * cm, 4.5 * cm, 2.5 * cm, 2.5 * cm, 2 * cm, 3.5 * cm],
        alinhar_direita={4, 5},
    )
    elementos.append(tabela_consolidado)
    elementos.append(PageBreak())

    for i, d in enumerate(desenhos):
        lista, versao, itens = d["lista"], d["versao"], d["itens"]
        elementos.append(Paragraph(f"Desenho {lista['numero_desenho']} — {lista['titulo'] or ''}", secao_estilo))
        elementos.append(Paragraph(
            f"Revisão: {versao['versao'] if versao else '-'} &nbsp;&nbsp;|&nbsp;&nbsp; "
            f"Nº Cliente: {lista['numero_cliente'] or '-'} &nbsp;&nbsp;|&nbsp;&nbsp; "
            f"Nº Fornecedor: {lista['numero_fornecedor'] or '-'}",
            sub_estilo,
        ))
        dados = [["Item", "Código", "Descrição", "Fabricante", "Bitola", "Quantidade", "Unidade"]]
        for idx, item in enumerate(itens, start=1):
            dados.append([idx, item["codigo"], item["descricao"], item["fabricante"] or "-",
                          item["bitola"] or "-", item["quantidade"], item["unidade"]])
        if not itens:
            dados.append(["-", "-", "Nenhum material nesta versão.", "-", "-", "-", "-"])
        tabela = _tabela_quebravel(
            dados,
            col_widths=[1.3 * cm, 2.8 * cm, 8 * cm, 4.5 * cm, 2.8 * cm, 2.8 * cm, 2.3 * cm],
            alinhar_direita={0, 5, 6},
        )
        elementos.append(tabela)
        if i < len(desenhos) - 1:
            elementos.append(PageBreak())

    if not desenhos:
        elementos.append(Paragraph("Nenhuma lista por desenho cadastrada neste projeto.", sub_estilo))

    doc.build(elementos)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f"relatorio_projeto_{projeto['codigo']}.pdf", mimetype="application/pdf")


def _carregar_projeto_para_relatorio(projeto_id):
    if not projeto_permitido(projeto_id):
        return None
    projeto = db.query_one(
        """SELECT p.*, c.razao_social AS cliente_nome, c.logo_url AS cliente_logo_url
           FROM projetos p LEFT JOIN clientes c ON c.id = p.cliente_id
           WHERE p.id = %s""",
        (projeto_id,),
    )
    if not projeto:
        return None
    config = {c["chave"]: c["valor"] for c in db.query_all("SELECT chave, valor FROM configuracoes")}
    return {"projeto": projeto, "config": config}


def _marca_dagua_preliminar(doc, ativa):
    """Callback onFirstPage/onLaterPages que carimba 'PRELIMINAR' na diagonal
    quando a versão exibida ainda é um rascunho (não emitida)."""
    def desenhar(canvas, _doc):
        if not ativa:
            return
        canvas.saveState()
        canvas.setFont("Helvetica-Bold", 70)
        canvas.setFillColor(colors.Color(0.85, 0.2, 0.2, alpha=0.15))
        canvas.translate(doc.pagesize[0] / 2, doc.pagesize[1] / 2)
        canvas.rotate(35)
        canvas.drawCentredString(0, 0, "PRELIMINAR")
        canvas.restoreState()
    return desenhar


def _carregar_versao_itens(projeto, versao_id_param, tipo):
    """Carrega a versão (a informada ou, por padrão, a última salva) e seus
    itens para relatórios de Lista PQ ou Lista de Compras. tipo é sempre um
    literal fixo no código ('pq' ou 'compras'), nunca vindo do usuário."""
    versao_id = versao_id_param or projeto[f"{tipo}_versao_atual_id"]
    if not versao_id:
        return None, []
    versao = db.query_one(
        f"""SELECT v.*, u.nome AS criado_por_nome FROM lista_{tipo}_versoes v
            LEFT JOIN usuarios u ON u.id = v.criado_por WHERE v.id = %s""",
        (versao_id,),
    )
    itens = db.query_all(
        f"""SELECT i.*, m.codigo, m.descricao, m.fabricante, m.bitola, m.unidade
            FROM lista_{tipo}_itens i JOIN materiais m ON m.id = i.material_id
            WHERE i.versao_id = %s ORDER BY m.codigo""",
        (versao_id,),
    )
    return versao, itens


# =========================================================
# RELATÓRIO — LISTA PQ
# =========================================================
@relatorios_bp.get("/api/projetos/<int:projeto_id>/lista-pq/relatorio/excel")
@login_required
def relatorio_lista_pq_excel(projeto_id):
    ctx = _carregar_projeto_para_relatorio(projeto_id)
    if not ctx:
        return jsonify({"erro": "Projeto não encontrado"}), 404
    projeto, config = ctx["projeto"], ctx["config"]
    versao, itens = _carregar_versao_itens(projeto, request.args.get("versao_id", type=int), "pq")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lista PQ"
    largura_total = 8
    fino = Side(style="thin", color="000000")
    borda = Border(left=fino, right=fino, top=fino, bottom=fino)
    centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    esquerda = Alignment(horizontal="left", vertical="center", wrap_text=True)

    data_versao = versao["criado_em"].strftime("%d/%m/%Y") if versao else "-"
    linha = _cabecalho_excel_com_logos(
        ws, largura_total, config.get("nome_empresa", ""), config.get("logo_url", ""),
        projeto["cliente_nome"], projeto.get("cliente_logo_url"),
        f"{projeto['codigo']} — {projeto['nome']}", "LISTA PQ",
        versao["versao"] if versao else "-", data_versao,
    )

    cabecalho = ["Item", "Código", "Descrição", "Fabricante", "Bitola", "Qtd. Base", "% Aplicado", "Qtd. Atualizada", "Unidade"]
    largura_total = len(cabecalho)
    for col, titulo in enumerate(cabecalho, start=1):
        cel = ws.cell(row=linha, column=col, value=titulo)
        cel.font = Font(bold=True, size=10, color="FFFFFF")
        cel.fill = openpyxl.styles.PatternFill("solid", fgColor="1f3a5f")
        cel.alignment, cel.border = centro, borda
    linha += 1
    for idx, item in enumerate(itens, start=1):
        valores = [idx, item["codigo"], item["descricao"], item["fabricante"] or "", item["bitola"] or "",
                   float(item["quantidade_base"]), f"{float(item['percentual']):g}%",
                   float(item["quantidade_atualizada"]), item["unidade"]]
        for col, valor in enumerate(valores, start=1):
            cel = ws.cell(row=linha, column=col, value=valor)
            cel.border = borda
            cel.alignment = esquerda if col == 3 else centro
        linha += 1
    if not itens:
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=largura_total)
        ws.cell(row=linha, column=1, value="Nenhuma versão salva da Lista PQ ainda.").alignment = centro

    for i, w in enumerate([6, 14, 32, 20, 12, 12, 12, 14, 10], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f"lista_pq_{projeto['codigo']}.xlsx",
                      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@relatorios_bp.get("/api/projetos/<int:projeto_id>/lista-pq/relatorio/pdf")
@login_required
def relatorio_lista_pq_pdf(projeto_id):
    ctx = _carregar_projeto_para_relatorio(projeto_id)
    if not ctx:
        return jsonify({"erro": "Projeto não encontrado"}), 404
    projeto, config = ctx["projeto"], ctx["config"]
    versao, itens = _carregar_versao_itens(projeto, request.args.get("versao_id", type=int), "pq")
    rascunho = bool(versao and versao["status"] == "rascunho")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=1.2 * cm, bottomMargin=1.2 * cm,
                             leftMargin=1.2 * cm, rightMargin=1.2 * cm)
    data_versao = versao["criado_em"].strftime("%d/%m/%Y") if versao else "-"
    elementos = _cabecalho_pdf_com_logos(
        config.get("nome_empresa", ""), config.get("logo_url", ""),
        projeto["cliente_nome"], projeto.get("cliente_logo_url"),
        f"{projeto['codigo']} — {projeto['nome']}", "LISTA PQ",
        versao["versao"] if versao else "-", data_versao,
    )

    dados = [["Item", "Código", "Descrição", "Fabricante", "Bitola", "Qtd. Base", "%", "Qtd. Atualizada", "Unidade"]]
    for idx, item in enumerate(itens, start=1):
        dados.append([idx, item["codigo"], item["descricao"], item["fabricante"] or "-", item["bitola"] or "-",
                      item["quantidade_base"], f"{float(item['percentual']):g}%", item["quantidade_atualizada"], item["unidade"]])
    if not itens:
        dados.append(["-", "-", "Nenhuma versão salva da Lista PQ ainda.", "-", "-", "-", "-", "-", "-"])

    tabela = _tabela_quebravel(
        dados, col_widths=[1.2 * cm, 2.5 * cm, 6.5 * cm, 3.5 * cm, 2.2 * cm, 2.2 * cm, 1.8 * cm, 2.8 * cm, 2 * cm],
        alinhar_direita={0, 5, 6, 7, 8},
    )
    elementos.append(tabela)
    marca_dagua = _marca_dagua_preliminar(doc, rascunho)
    doc.build(elementos, onFirstPage=marca_dagua, onLaterPages=marca_dagua)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f"lista_pq_{projeto['codigo']}.pdf", mimetype="application/pdf")


# =========================================================
# RELATÓRIO — LISTA DE COMPRAS
# =========================================================
@relatorios_bp.get("/api/projetos/<int:projeto_id>/lista-compras/relatorio/excel")
@login_required
def relatorio_lista_compras_excel(projeto_id):
    ctx = _carregar_projeto_para_relatorio(projeto_id)
    if not ctx:
        return jsonify({"erro": "Projeto não encontrado"}), 404
    projeto, config = ctx["projeto"], ctx["config"]
    versao, itens = _carregar_versao_itens(projeto, request.args.get("versao_id", type=int), "compras")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lista de Compras"
    fino = Side(style="thin", color="000000")
    borda = Border(left=fino, right=fino, top=fino, bottom=fino)
    centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    esquerda = Alignment(horizontal="left", vertical="center", wrap_text=True)

    data_versao = versao["criado_em"].strftime("%d/%m/%Y") if versao else "-"
    linha = _cabecalho_excel_com_logos(
        ws, 6, config.get("nome_empresa", ""), config.get("logo_url", ""),
        projeto["cliente_nome"], projeto.get("cliente_logo_url"),
        f"{projeto['codigo']} — {projeto['nome']}", "LISTA DE COMPRAS",
        versao["versao"] if versao else "-", data_versao,
    )

    cabecalho = ["Item", "Código", "Descrição", "Fabricante", "Bitola", "Quantidade", "Unidade"]
    largura_total = len(cabecalho)
    for col, titulo in enumerate(cabecalho, start=1):
        cel = ws.cell(row=linha, column=col, value=titulo)
        cel.font = Font(bold=True, size=10, color="FFFFFF")
        cel.fill = openpyxl.styles.PatternFill("solid", fgColor="1f3a5f")
        cel.alignment, cel.border = centro, borda
    linha += 1
    for idx, item in enumerate(itens, start=1):
        valores = [idx, item["codigo"], item["descricao"], item["fabricante"] or "", item["bitola"] or "",
                   float(item["quantidade"]), item["unidade"]]
        for col, valor in enumerate(valores, start=1):
            cel = ws.cell(row=linha, column=col, value=valor)
            cel.border = borda
            cel.alignment = esquerda if col == 3 else centro
        linha += 1
    if not itens:
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=largura_total)
        ws.cell(row=linha, column=1, value="Nenhuma versão salva da Lista de Compras ainda.").alignment = centro

    for i, w in enumerate([6, 14, 34, 20, 12, 12, 10], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f"lista_compras_{projeto['codigo']}.xlsx",
                      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@relatorios_bp.get("/api/projetos/<int:projeto_id>/lista-compras/relatorio/pdf")
@login_required
def relatorio_lista_compras_pdf(projeto_id):
    ctx = _carregar_projeto_para_relatorio(projeto_id)
    if not ctx:
        return jsonify({"erro": "Projeto não encontrado"}), 404
    projeto, config = ctx["projeto"], ctx["config"]
    versao, itens = _carregar_versao_itens(projeto, request.args.get("versao_id", type=int), "compras")
    rascunho = bool(versao and versao["status"] == "rascunho")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=1.2 * cm, bottomMargin=1.2 * cm,
                             leftMargin=1.2 * cm, rightMargin=1.2 * cm)
    data_versao = versao["criado_em"].strftime("%d/%m/%Y") if versao else "-"
    elementos = _cabecalho_pdf_com_logos(
        config.get("nome_empresa", ""), config.get("logo_url", ""),
        projeto["cliente_nome"], projeto.get("cliente_logo_url"),
        f"{projeto['codigo']} — {projeto['nome']}", "LISTA DE COMPRAS",
        versao["versao"] if versao else "-", data_versao,
    )

    dados = [["Item", "Código", "Descrição", "Fabricante", "Bitola", "Quantidade", "Unidade"]]
    for idx, item in enumerate(itens, start=1):
        dados.append([idx, item["codigo"], item["descricao"], item["fabricante"] or "-", item["bitola"] or "-",
                      item["quantidade"], item["unidade"]])
    if not itens:
        dados.append(["-", "-", "Nenhuma versão salva da Lista de Compras ainda.", "-", "-", "-", "-"])

    tabela = _tabela_quebravel(
        dados, col_widths=[1.3 * cm, 2.8 * cm, 8 * cm, 4.5 * cm, 2.8 * cm, 2.8 * cm, 2.3 * cm],
        alinhar_direita={0, 5, 6},
    )
    elementos.append(tabela)
    marca_dagua = _marca_dagua_preliminar(doc, rascunho)
    doc.build(elementos, onFirstPage=marca_dagua, onLaterPages=marca_dagua)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f"lista_compras_{projeto['codigo']}.pdf", mimetype="application/pdf")
